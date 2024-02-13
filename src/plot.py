# -*- coding: utf-8 -*-
"""
Title: Psychrometric Chart
Author: Alexander Weaver
Co-Authors: Virginia Covert, Korynn Haetten, Stanley Moonjeli

Description:
    A script to generate the psychrometric charts and mass vs. time
    plots for the DesiGators IPPD project.

Sponsor: Dr. Andrew MacIntosh
Coach: Dr. Philip Jackson

Thanks:
    Prathamesh Nachane - for writing a great deal of CoolProp code
        which is borrowed from herein.

References:
    1 - https://github.com/prathamesh-nachane/Psychometric-chart-in-Python/blob/master/psycochart.py
"""

import numpy as np
import matplotlib.pyplot as plt
import openpyxl
import pandas as pd
import time
import os
from CoolProp.HumidAirProp import HAPropsSI


# input array is _ height x 24 width with 1st column = time and 1st row = headers


def csv_load(in_fold, filename):
    # load csv files necessary to produce psychrometric and mass plots into arrays

    pd.read_csv(in_fold + filename + '.csv', delimiter=",").to_excel(in_fold + filename + '.xlsx', index=False)

    wb = openpyxl.load_workbook(in_fold + filename + '.xlsx',
                                data_only=True)  # insert workbook name here. data_only=True ignores formulas
    ws = wb.active  # locates sheet to be read

    i = 2  # initial row count

    # mass_data = ['Time', '1', '2', '3', '4']
    # rht_data = ['Time', '1I T', '1O T', '2I T', '2O T', '3I T', '3O T', '4I T', '4O T',
    #            '1I RH', '1O RH', '2I RH', '2O RH', '3I RH', '3O RH', '4I RH', '4O RH']

    mass_data = np.zeros(5)
    rht_data = np.zeros(17)

    while str(ws.cell(row=i, column=1).value) != 'None':
        # terminates addition to array when no data is found
        # initialize list for current row to append to growing array

        list_mass = []
        list_rht = []

        # add time value to first column in each row

        list_mass = np.append(list_mass, ws.cell(row=i, column=1).value)
        list_rht = np.append(list_rht, ws.cell(row=i, column=1).value)

        for k in np.arange(2, 10, 2):
            mass_tot = float(ws.cell(row=i, column=k).value + ws.cell(row=i, column=k + 1).value)
            list_mass = np.append(list_mass, mass_tot)

        mass_data = np.vstack((mass_data, list_mass))

        for k in np.arange(10, 26, 1):
            list_rht = np.append(list_rht, float(ws.cell(row=i, column=k).value))

        rht_data = np.vstack((rht_data, list_rht))

        i += 1

    return mass_data, rht_data


def mass_plot(mass_points, points_interval):
    print(mass_points)

    mass_points = np.delete(mass_points, 0, 0)

    mass_points = np.array(mass_points)

    # plot subset of data points to reduce graph clutter if necessary

    mass_points_new = mass_points[0][:]

    for i in np.arange(points_interval, len(mass_points), points_interval):
        mass_points_new = np.vstack((mass_points_new, mass_points[i]))

    print(mass_points_new)

    plt.figure(1)

    # plotting data for each chamber

    plt.plot(mass_points_new[:, 0], mass_points_new[:, 1], label='Chamber A')
    plt.plot(mass_points_new[:, 0], mass_points_new[:, 2], label='Chamber B')
    plt.plot(mass_points_new[:, 0], mass_points_new[:, 3], label='Chamber C')
    plt.plot(mass_points_new[:, 0], mass_points_new[:, 4], label='Chamber D')

    plt.legend()
    plt.xlabel('Time [min]')
    plt.ylabel('Food mass [kg]')

    # saving plot as image

    path = r'C:\Users\benco\OneDrive\Desktop\DesiGators\src'  # add whatever the path is

    name = str(time.time())

    new_path = path + r'/plots'

    if not os.path.exists(new_path):
        os.makedirs(new_path)

    plt.savefig('plots/' + name + '.jpg')

    imgname = 'plots/' + name

    return imgname


def plot_psy_chart(x_low_limit=20, x_upp_limit=60, y_low_limit=0, y_upp_limit=0.03, p=101325, RH_lines='y',
                   H_lines='y', WB_lines='y'):
    Tdb = np.linspace(x_low_limit, x_upp_limit, 100) + 273.15

    # Make the figure and the axes
    fig, ax = plt.subplots(figsize=(10, 8))
    fig.patch.set_alpha(0.9)
    ax.plot()
    ax.set_xlim(Tdb[0] - 273.15, Tdb[-1] - 273.15)
    ax.set_ylim(y_low_limit, y_upp_limit)
    ax.set_xlabel(r"$T_{db}$ [$^{\circ}$C]")
    ax.set_ylabel(r"$Abs. humidity$ ($m_{w}/m_{da}$) [-]")
    ax.xaxis.label.set_fontsize(15)
    ax.yaxis.label.set_fontsize(15)
    ax.tick_params(axis='x', labelsize=15)
    ax.tick_params(axis='y', labelsize=15)

    # Saturation line
    w = [HAPropsSI('W', 'T', T, 'P', p, 'R', 1.0) for T in Tdb]
    ax.plot(Tdb - 273.15, w, lw=2)

    # Enthalpy lines
    if H_lines == 'y':
        H_lines = [-20000, -10000, 0, 10000, 20000, 30000, 40000, 50000, 60000, 70000, 80000, 90000]
        for H in H_lines:
            # Line goes from saturation to zero humidity ratio for this enthalpy
            T1 = HAPropsSI('T', 'H', H, 'P', p, 'R', 1.0) - 273.15
            T0 = HAPropsSI('T', 'H', H, 'P', p, 'R', 0.0) - 273.15
            w1 = HAPropsSI('W', 'H', H, 'P', p, 'R', 1.0)
            w0 = HAPropsSI('W', 'H', H, 'P', p, 'R', 0.0)
            ax.plot(np.r_[T1, T0], np.r_[w1, w0], 'go--', lw=1, alpha=0.5)
            string = r'$H$=' + '{s:0.0f}'.format(s=H / 1000) + ' kJ/kg'
            bbox_opts = dict(boxstyle='square,pad=0.0', fc='white', ec='None', alpha=0)
            ax.text(T1 - 2, w1 + 0.0005, string, size='small', ha='center', va='center', bbox=bbox_opts)

    # Wet-bulb temperature lines
    if WB_lines == 'y':
        WB_lines = np.linspace(0, 55, 12) + 273.15
        for WB in WB_lines:
            # Line goes from saturation to zero humidity ratio for this enthalpy
            T1 = HAPropsSI('T', 'Twb', WB, 'P', p, 'R', 1.0) - 273.15 - 2
            T0 = HAPropsSI('T', 'Twb', int(WB), 'P', int(p), 'R', 0) - 273.15
            wb1 = HAPropsSI('W', 'Twb', WB, 'P', p, 'R', 1) + 0.002
            wb0 = HAPropsSI('W', 'Twb', int(WB), 'P', int(p), 'R', 0.0)
            ax.plot(np.r_[T1, T0], np.r_[wb1, wb0], 'm--', lw=1, alpha=0.5)
            string = r'$WB$=' + '{s:0.0f}'.format(s=(WB - 273)) + ' [C]'
            bbox_opts = dict(boxstyle='square,pad=0.0', fc='white', ec='None', alpha=0)
            ax.text(T1 - 2, wb1 + 0.0005, string, size='small', ha='center', va='center', bbox=bbox_opts)

    # Humidity lines
    if RH_lines == 'y':
        RH_lines = [0.05, 0.1, 0.15, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9]
        for RH in RH_lines:
            w = [HAPropsSI('W', 'T', T, 'P', p, 'R', RH) for T in Tdb]
            ax.plot(Tdb - 273.15, w, 'r--', lw=1, alpha=0.5)
            yv = [HAPropsSI('W', 'T', T, 'P', p, 'R', RH) for T in Tdb]
            T_K = Tdb[round(95.4082 - 40.8163 * RH)]
            w = yv[round(95.4082 - 40.8163 * RH)]
            string = r'$\phi$=' + '{s:0.0f}'.format(s=RH * 100) + '%'
            bbox_opts = dict(boxstyle='square,pad=0.0', fc='white', ec='None', alpha=0)
            ax.text(T_K - 273.15, w, string, size='medium', ha='center', va='center', bbox=bbox_opts)
    #    plt.close('all')
    return (fig, ax)


# %% code to overlay points
def plot_points(arr, figure, axes, col='b', typ='-', grid='on'):
    colstr = col + 'o' + typ
    b = np.zeros(len(arr))
    c = np.zeros(len(arr))
    for i in range(len(arr)):
        b[i] = arr[i][0]
        c[i] = arr[i][1]
        # label = str(calc_prop_of(i, arr[i][0], arr[i][1]))
        # axes.scatter(b[i], c[i], s=30, color=col, label=label)
        # axes.legend(loc=0, fontsize='xx-small', framealpha=0.25)
    axes.plot(b, c, colstr)
    if grid == 'on':
        axes.grid(linestyle='--', alpha=0.5, linewidth=1)
    # for i in range(len(arr)):
    # axes.text(1.01 * arr[i][0], 1.05 * arr[i][1], str(i + 1), style='italic', size='medium',
    # bbox={'facecolor': col, 'alpha': 0.5, 'pad': 1})
    axes.plot()
    return (figure, axes)


def calc_prop_of(counter, xdata, ydata):
    a = 'Point: ' + str(counter + 1)
    b = "-- R = " + str(round(100 * HAPropsSI('R', 'T', xdata + 273, 'P', 101325, 'W', ydata), 2)) + ' %'
    c = '-- T = ' + str(round(xdata, 2)) + ' [C]'
    d = '-- W = ' + str(round(ydata, 4))
    e = '-- H = ' + str(round((HAPropsSI('H', 'T', xdata + 273, 'P', 101325, 'W', ydata) / 1000), 3)) + ' kJ/kg'
    #       f =' W = '+ str(100*HAPropsSI('Twb','T',xdata+273,'P',101325,'W',ydata)) +' [C]'
    return (str(a + b + c + d + e))


def plot_psy_chart_w_points(psychro_points):
    plt.figure(2)

    # plt.close("all")

    figure, axes = plot_psy_chart(x_low_limit=-10, x_upp_limit=60, y_low_limit=0, y_upp_limit=0.03, p=101325,
                                  RH_lines='y',
                                  H_lines='y', WB_lines='y')

    # psychro_points will take the place of variable 'a'

    figure, axes = plot_points(psychro_points, figure, axes, col='r', typ='-', grid='on')

    # saving plot as image

    path = r'C:\Users\benco\OneDrive\Desktop\DesiGators\src'  # add whatever the path is

    name = str(time.time())

    new_path = path + '/Psychrometric Plots'

    if not os.path.exists(new_path):
        os.makedirs(new_path)

    plt.savefig('Psychrometric Plots/' + name + '.jpg')

    imgname = 'Psychrometric Plots/' + name

    plt.show()

    return imgname


def main():
    mass_vals, rht_vals = csv_load('', 'datatest')

    mass_plot(mass_vals, 4)
    plt.show()

    psychro_points = [[50, 0.007], [40, 0.006], [30, 0.003]]
    plot_psy_chart_w_points(psychro_points)


if __name__ == '__main__':
    main()
